# Python script to utilize PLEXOS CLI
# 8-Dec-2023
# replacing API commands with CLI
# StudyID is 9e9f66ca-9fab-4154-b356-0916e014681d


import os, sys, re, clr
import subprocess as sp
import pandas as pd
from datetime import datetime
from openpyxl import *
import openpyxl
import builtins
import xlwings # update formula in excel sheet
import json #handling json config and messages
import time

from shutil import copyfile

plexospath = 'C:/Program Files/Energy Exemplar/PLEXOS 9.2 API'
sys.path.append(plexospath)
clr.AddReference('PLEXOS_NET.Core')
clr.AddReference('EEUTILITY')
clr.AddReference('EnergyExemplar.PLEXOS.Utility')

# .NET related imports
from PLEXOS_NET.Core import DatabaseCore, Solution, PLEXOSConnect
from EEUTILITY.Enums import *
from EnergyExemplar.PLEXOS.Utility.Enums import *
from System import String


def update_dataset(original_ds, new_tariff): #, new_ds
    if os.path.exists(original_ds):

        # # delete the modified file if it already exists
        # if os.path.exists(new_ds):
        #     os.remove(new_ds)

        # # copy the PLEXOS input file
        # copyfile(original_ds, new_ds)
        
        # Create an object to store the input data
        db = DatabaseCore()
        db.Connection(original_ds)
        
        # # Add a category
        # '''
        # Int32 AddCategory(
        #     ClassEnum nClassId,
        #     String strCategory
        #     )
        # '''
        # db.AddCategory(ClassEnum.Generator, 'API')
        
        # # Add an object (and the System Membership)
        # '''
        # Int32 AddObject(
        #     String strName,
        #     ClassEnum nClassId,
        #     Boolean bAddSystemMembership,
        #     String strCategory[ = None],
        #     String strDescription[ = None]
        #     )
        # '''
        # db.AddObject('ApiGen', ClassEnum.Generator, True, 'API', 'Testing the API')
        
        # # Add memberships
        # '''
        # Int32 AddMembership(
        #     CollectionEnum nCollectionId,
        #     String strParent,
        #     String strChild
        #     )
        # '''
        # db.AddMembership(CollectionEnum.GeneratorNodes, 'ApiGen', '101')    
        # db.AddMembership(CollectionEnum.GeneratorFuels, 'ApiGen', 'Coal/Steam')
        # db.AddMembership(CollectionEnum.ReserveGenerators, 'Spin Up', 'ApiGen')
        
        # Get the System.Generators membership ID for this new generator
        '''
        Int32 GetMembershipID(
            CollectionEnum nCollectionId,
            String strParent,
            String strChild
            )    
        '''
        mem_id = db.GetMembershipID(CollectionEnum.SystemConstraints, \
                                    'System', 'Time-Block DFR')
        # print("Constraint ID is ", mem_id)                                    
                                    
        # Add properties
        '''
        Int32 AddProperty(
            Int32 MembershipId,
            Int32 EnumId,
            Int32 BandId,
            Double Value,
            Object DateFrom,
            Object DateTo,
            Object Variable,
            Object DataFile,
            Object Pattern,
            Object Scenario,
            Object Action,
            PeriodEnum PeriodTypeId
            )
        
        Also we need to obtain the EnumId for each property
        that we intend to add
        '''
        # Remove property
        flag = db.RemoveProperty(mem_id, int(SystemConstraintsEnum.PenaltyPrice), \
                1, None, None, '1.5 x Scalar', None, None, 'Penalty Price', \
                1, PeriodEnum.Interval)

        # Add a property
        flag  += db.AddProperty(mem_id, int(SystemConstraintsEnum.PenaltyPrice), \
                1, new_tariff, None, None, '1.5 x Scalar', None, None, 'Penalty Price', \
                1, PeriodEnum.Interval)
        
        if flag == 2:
            print("A. Property replacement complete\n")
        
        # save the data set
        db.Close()

        # push changes to cloud
        sp.call('plexos-cloud.exe study changeset push --studyId 9e9f66ca-9fab-4154-b356-0916e014681d --message "Update Tariff"', shell=True)

        # get latest changeset ID
        # cs = sp.call('pxc study changeset get-latest --studyId 9e9f66ca-9fab-4154-b356-0916e014681d', shell=True)
        res = sp.check_output('pxc study changeset get-latest --studyId 9e9f66ca-9fab-4154-b356-0916e014681d', shell=True)
        cs = res.decode("utf-8")
        print("Updated changeset is ", cs)
        return cs

					
#Execute the model 
def run_model(plexospath, filename, foldername, modelname):
# def run_model(plexospath, filename, foldername, modelname, username, password): #with credentials
    # launch the model on the local desktop
    # The \n argument is very important because it allows the PLEXOS
    # engine to terminate after completing the simulation
    print("B. Run the model")
    sp.call([os.path.join(plexospath, 'PLEXOS64.exe'), filename, r'\n', r'\o', foldername, r'\m', modelname])
    # sp.call([os.path.join(plexospath, 'PLEXOS64.exe'), filename, r'\n', r'\o', foldername, r'\m', modelname, r'\cu', username, r'\cp', password]) #with username
    

def parse_logfile(pattern, foldername, modelname, linecount = 1):
    
    currentlines = 0
    lines = []
    regex = re.compile(pattern)
    
    for line in builtins.open(os.path.join(foldername, 'Model ( {} ) Log.txt'.format(modelname))):
        if len(regex.findall(line)) > 0:
            currentlines = linecount
            
        if currentlines > 0:
            lines.append(line)
            currentlines -= 1
            
        if currentlines == 0 and len(lines) > 0:
            retval = '\n'.join(lines)
            lines = []
            yield retval

def query_results(sol_file, xlsx_file):
    #Query Results
    # Create a PLEXOS solution file object and load the solution
    if not os.path.exists(sol_file):
        print('No such file')
        return
        
    sol = Solution()
    sol.Connection(sol_file)
    sol.DisplayAlerts = False

    # print("Update EXCEL sheet with new results")

    # # connect to excel
    # book = load_workbook(xlsx_file)
    writer = pd.ExcelWriter(xlsx_file, mode='a', if_sheet_exists='overlay', engine='openpyxl')

    '''
    Simple query: works similarly to PLEXOS Solution Viewer

    Solution.Query(phase, collection, parent, child, period, series, props)
        phase -> SimulationPhaseEnum
        collection -> CollectionEnum
        parent -> the name of a parent object or ''
        child -> the name of a child object or ''
        period -> PeriodEnum
        series -> SeriesTypeEnum
        props -> a string containing an integer indicating the Property to query or ''
    returns a ADODB recordset... however, you don't *need* to worry about that...
    '''

    '''
    Simple query: works similarly to PLEXOS Solution Viewer
    
    QueryToList(
    	SimulationPhaseEnum SimulationPhaseId,
    	CollectionEnum CollectionId,
    	String ParentName,
    	String ChildName,
    	PeriodEnum PeriodTypeId,
    	SeriesTypeEnum SeriesTypeId,
    	String PropertyList[ = None],
    	Object DateFrom[ = None],
    	Object DateTo[ = None],
    	String TimesliceList[ = None],
    	String SampleList[ = None],
    	String ModelName[ = None],
    	AggregationEnum AggregationType[ = None],
    	String Category[ = None],
    	String Filter[ = None]
    	)
    '''    
    #AnnualizedBuildCost

    # To list all the enabled properties that can be reported
    # ReportObj=sol.GetReportedProperties()
    # print(list(ReportObj))

    #Limit the columns in the output
    # sol.SetSelectedColumns(['child_name','property_name','value'])

    #to query Battery properties
    propId = sol.PropertyName2EnumId("System", "Battery", "Batteries", "Annualized Build Cost")
    # print("BESS Annualized Cost")
    # print(propId)

    listresultsbat = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemBatteries, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              '67')

    # Check to see if the query had results
    if listresultsbat is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["BESS"]
        df_batAC = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsbat], columns=columns)
        df_output = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsbat], columns=columns)
        
        #save to excel
        df_batAC.to_excel(writer, "WACC", header=None, index=False, startcol=6,startrow=1)

    propId = sol.PropertyName2EnumId("System", "Battery", "Batteries", "Units Built")
    # print("BESS Units")
    # print(propId)
    # battery units
    listresultsbatu = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemBatteries, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId))

    # Check to see if the query had results
    if listresultsbatu is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["BESS"]
        df_batu = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsbatu], columns=columns)
        
        #save to excel
        df_batu.to_excel(writer, "WACC", header=None, index=False, startcol=9,startrow=28)

    #to query Battery - FOM properties
    propId = sol.PropertyName2EnumId("System", "Battery", "Batteries", "FO&M Cost")
    # print("BESS FO&M Cost")
    # print(propId)

    listresultsbatf = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemBatteries, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId))

    # Check to see if the query had results
    if listresultsbatf is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["BESS"]
        df_batf = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsbatf], columns=columns)
        
        #save to excel
        df_batf.to_excel(writer, "WACC", header=None, index=False, startcol=14,startrow=1)

    #to query Generation properties
    # annualized build cost 
    propId = sol.PropertyName2EnumId("System", "Generator", "Generators", "Annualized Build Cost")
    # print("Generator Annualized Cost")
    # print(propId)
    # propId2 = sol.PropertyName2EnumId("System", "Generator", "Generators", "FO&M Cost")
    # print(propId2)

    listresultsGAC = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemGenerators, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId))

    # Check to see if the query had results
    if listresultsGAC is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["Solar", "Wind"]
        dfGAC = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsGAC], columns=columns)
        
        #save to excel
        dfGAC.to_excel(writer, "WACC", header=None, index=False, startcol=4,startrow=1)

    propId = sol.PropertyName2EnumId("System", "Generator", "Generators", "Units Built")
    # print("Gens Units")
    # print(propId)
    # battery units
    listresultGenu = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemGenerators, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId))

    # Check to see if the query had results
    if listresultGenu is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["Solar", "Wind"]
        df_genu = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultGenu], columns=columns)
        
        #save to excel
        df_genu.to_excel(writer, "WACC", header=None, index=False, startcol=7,startrow=28)

    # to be closed once
    # writer._save()

    # FO&M cost
    propId2 = sol.PropertyName2EnumId("System", "Generator", "Generators", "FO&M Cost")
    # print("Generator FO&M Cost") 
    # print(propId2)

    listresultsGOM = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemGenerators, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId2))

    # Check to see if the query had results
    if listresultsGOM is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["Solar", "Wind"]
        dfGOM = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsGOM], columns=columns)
        
        dfGOM.to_excel(writer, "WACC", header=None, index=False, startcol=12,startrow=1)


    # Collect market sales
    # sales
    propId3 = sol.PropertyName2EnumId("System", "Market", "Markets", "Sales")
    # print("Market Sales")
    # print(propId3)

    listresultsMS = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemMarkets, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId3))

    # Check to see if the query had results
    if listresultsMS is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["FDRE", "IEX Sales"]
        dfMS = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsMS], columns=columns)
        
        dfMS["FDRE"].to_excel(writer, "WACC", header=None, index=False, startcol=4,startrow=28)  
        dfMS["IEX Sales"].to_excel(writer, "WACC", header=None, index=False, startcol=20,startrow=1)


    # Collect market purchases
    # purchases
    propId3 = sol.PropertyName2EnumId("System", "Market", "Markets", "Purchases")
    # print("Market Purchases")
    # print(propId3)

    listresultsMP = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemMarkets, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId3))

    # Check to see if the query had results
    if listresultsMP is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["IEX Purchases"]
        dfMP = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsMP], columns=columns)
        dfMP.to_excel(writer, "WACC", header=None, index=False, startcol=21,startrow=1)
          

    #purchase cost directly
    propId3 = sol.PropertyName2EnumId("System", "Market", "Markets", "Total Cost")
    # print("Market Purchase Cost") 
    # print(propId3)

    listresultsMPC = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemMarkets, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId3))

    # Check to see if the query had results
    if listresultsMPC is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["IEX Purchases"]
        dfMPC = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsMPC], columns=columns)
        dfMPC.to_excel(writer, "WACC", header=None, index=False, startcol=23,startrow=1)
           

    # Constraint results
    # violation and penalty costs
    propId3 = sol.PropertyName2EnumId("System", "Constraint", "Constraints", "Penalty Cost")
    # print("Constraint Penalty Cost") 
    # print(propId3)

    listresultsCP = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemConstraints, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId3))

    # Check to see if the query had results
    if listresultsCP is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["Time-block DFR"]
        dfCP = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsCP], columns=columns)
        # print(dfCP)
        
        dfCP.to_excel(writer, "WACC", header=None, index=False, startcol=13,startrow=28)
        
    propId3 = sol.PropertyName2EnumId("System", "Constraint", "Constraints", "Violation")
    # print("Constraint Violation") 
    # print(propId3)

    listresultsCV = sol.QueryToList(SimulationPhaseEnum.LTPlan, \
              CollectionEnum.SystemConstraints, \
              '', \
              '', \
              PeriodEnum.FiscalYear, \
              SeriesTypeEnum.Names, \
              str(propId3))

    # Check to see if the query had results
    if listresultsCV is None:
        print('No results')
    else:
        # Create a DataFrame with a column for each column in the results
        columns = ["Time-block DFR"]
        dfCV = pd.DataFrame([[row.GetProperty.Overloads[String](n) for n in columns] for row in listresultsCV], columns=columns)
        # print(dfCP)
        
        dfCV.to_excel(writer, "WACC", header=None, index=False, startcol=12,startrow=28)

    # close connection to excel
    # writer._save() 
    writer.close()
    writer.handles = None

    # update formulas
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(xlsx_file)
    excel_book.save()
    excel_book.close()
    excel_app.quit()

    print("C. EXCEL sheet is updated")  

# def read_value(xlsx_file):
#     tariff = 0
#     # connect to excel sheet
#     wb = openpyxl.load_workbook(xlsx_file, data_only=True) #, data_only=True read only the calculated value
#     sh = wb.active
#     #Using cell() function
#     c3 = sh.cell(row=29,column=23)
#     tariff = 1000 * c3.value
#     wb.close()
    
#     print("A. Read the tariff value ", tariff)
#     return tariff

def check_convg(LCOE_val, xlsx_file):
    # connect to excel sheet
    wb = openpyxl.load_workbook(xlsx_file, data_only=True) #, data_only=True read only the calculated value
    sh = wb.active
    #Using cell() function
    LCOE_val_n = sh.cell(row=29,column=23)
    test = False
    change = 100*(LCOE_val_n.value - LCOE_val)/LCOE_val
    if (abs(change) <= 0.5):
        print("Convergence reached")
        test = True
    wb.close()
    
    print("D. New/old values of LCOE",LCOE_val_n.value,", ",LCOE_val)
    return test,LCOE_val_n.value

def main():
    plexospath = "C:\Program Files\Energy Exemplar\PLEXOS 9.2"
    count = 0
    PLEXOS_db = 'FDRE_IV_Model_Auto.xml'
    Excel_sheet = 'Yearly Peak 1 80 Cloud.xlsx'
    LCOE_val = 4300 # initial value
    tariff_list = []
    tariff_list.append(LCOE_val)
    study_changeset = None
    sim_id = '2f1802a4-9f1d-454f-9627-886184ce7ab3' #from a successful run
    config_file = "runConfig.json"

    #login to PLEXOS Cloud
    study_changeset = connect_PXC()

    while count < 6:    

        # update tariff value for penalty cost
        study_changeset = update_dataset(PLEXOS_db, LCOE_val*1000)

        # update the changeset ID in runConfig.json
        with builtins.open("runConfig.json", "r") as jsonFile:
            data = json.load(jsonFile)

        # update the field
        data["ChangeSetId"] = study_changeset.strip() # remove extra /n

        with builtins.open("runConfig.json", "w") as jsonFile:
            json.dump(data, jsonFile, indent=3)

        # Execute model
        # run_model(plexospath, PLEXOS_db, '.', '1. LT_Cost_Optimal_90') # can add the username and password at the end
        print("B. Run the model")
        res = sp.check_output('pxc simulation enqueue --file runConfig.json', shell=True)
        msg = res.decode("utf-8")
        print("Simulation enqueued as",msg)

        # extract simulation ID
        msgjson = json.loads(msg)
        sim_id = msgjson[0]["Id"] # first 0 is list index, next is the dictionary index
        print("Initiated simulation ID is ", sim_id)

        # wait for simulation run to complete        
        while (True):
            text1 = 'pxc simulation progress --simulationId '+sim_id
            # res = sp.check_output('pxc simulation progress --simulationId 2f1802a4-9f1d-454f-9627-886184ce7ab3 -aut', shell=True)

            msg2 = sp.run(text1, shell=True, capture_output=True, text=True)
            print("Current status is ",msg2.stdout)
            run_status = msg2.stdout

            if (run_status.strip() == "Status: Completed Success"):
                print("Simulation complete")
                break
            else:
                print("waiting for simulation to complete")
                time.sleep(120) # Makes Python wait for 10 sec        
              

        # identify solutionId from simulationID
        text1 = "pxc solution list --simulationId "+sim_id+" --type Raw -aut"
        res = sp.check_output(text1, shell=True)
        msg2 = res.decode("utf-8")

        # # extract solution ID
        msg2json = json.loads(msg2)
        sol_id = msg2json["EventData"]["Solutions"][0]["SolutionId"]["Value"] # first 0 is list index, next is the dictionary index
        print("Solution ID is ", sol_id)

        # # download results or artifacts
        text1 = "pxc solution files download --solutionId "+sol_id+" --type Raw --outputDirectory .\CloudResults --overwrite"
        res = sp.check_output(text1, shell=True)
        msg2 = res.decode("utf-8")

        text1 = "pxc solution files download --solutionId "+sol_id+" --type Log --outputDirectory .\CloudResults --overwrite"
        res = sp.check_output(text1, shell=True)
        msg2 = res.decode("utf-8")

        # Check log file
        for res in parse_logfile('LT Plan Completed','.\\CloudResults','1. LT_Cost_Optimal_90',25):
            print(res)

        # update result to excel sheet
        query_results('CloudResults/Model 1. LT_Cost_Optimal_90 Solution.zip', Excel_sheet)   

        # Check for convergence
        convg, new_lcoe = check_convg(LCOE_val, Excel_sheet)
        LCOE_val = new_lcoe
        tariff_list.append(new_lcoe)
        if convg == True:
            break

        count += 1
        print("Iter no: ",count) 

    print("Tariff list: ", tariff_list)

def connect_PXC():
    # login 
    # subprocess.call('pxc auth login', shell=True) #uncomment this line to login

    # check login status
    sp.call('pxc auth status', shell=True)

    # check latest changeset
    # cs = sp.call('pxc study changeset get-latest --studyId 9e9f66ca-9fab-4154-b356-0916e014681d', shell=True)
    res = sp.check_output('pxc study changeset get-latest --studyId 9e9f66ca-9fab-4154-b356-0916e014681d', shell=True)
    cs = res.decode("utf-8")  
    print("Initial changeset is ", cs) #bytes to string
    return cs

# call main function
if __name__ == '__main__':
    main()


